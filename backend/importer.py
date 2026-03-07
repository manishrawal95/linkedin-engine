"""
LinkedIn Creator Analytics XLSX importer.

Parses all 5 sheets from LinkedIn's export:
  DISCOVERY — overall impressions, members reached
  ENGAGEMENT — daily impressions + engagements
  TOP POSTS — post URLs with engagement counts and impressions
  FOLLOWERS — daily new follower counts + total
  DEMOGRAPHICS — job titles, locations, industries, seniority, company size
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone

from backend.db import get_conn

logger = logging.getLogger(__name__)


def _safe_int(val: object) -> int:
    if val is None or val == "":
        return 0
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return 0


def _safe_float(val: object) -> float:
    if val is None or val == "":
        return 0.0
    try:
        cleaned = str(val).replace(",", "").replace("%", "").strip()
        if cleaned.startswith("< "):
            return 0.005  # "< 1%" → 0.5%
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def _parse_date(val: object) -> str | None:
    """Parse m/d/yyyy or other date formats to YYYY-MM-DD."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Try parsing datetime objects from openpyxl
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return None


def _extract_activity_id(url: str) -> str | None:
    """Extract the numeric activity ID from any LinkedIn post URL format.

    Handles:
      - /posts/activity-7429928090975428609-_lAQ?...
      - /feed/update/urn:li:activity:7429928090975428609
      - /posts/username_...-activity-7429928090975428609-...
    """
    import re

    m = re.search(r"activity[:\-](\d{19,20})", url)
    return m.group(1) if m else None


def import_linkedin_xlsx(file_bytes: bytes) -> dict:
    """Parse and import a full LinkedIn Creator Analytics XLSX export.

    Returns summary dict with counts for each sheet processed.
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError(
            "XLSX support requires openpyxl. Install with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = {s.upper(): s for s in wb.sheetnames}
    batch_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    result: dict = {"batch_id": batch_id, "sheets_processed": []}

    # Process each sheet
    if "DISCOVERY" in sheets:
        r = _import_discovery(wb[sheets["DISCOVERY"]], batch_id)
        result["discovery"] = r
        result["sheets_processed"].append("DISCOVERY")

    if "ENGAGEMENT" in sheets:
        r = _import_engagement(wb[sheets["ENGAGEMENT"]], batch_id)
        result["engagement"] = r
        result["sheets_processed"].append("ENGAGEMENT")

    if "TOP POSTS" in sheets:
        r = _import_top_posts(wb[sheets["TOP POSTS"]], batch_id)
        result["top_posts"] = r
        result["sheets_processed"].append("TOP POSTS")

    if "FOLLOWERS" in sheets:
        r = _import_followers(wb[sheets["FOLLOWERS"]], batch_id)
        result["followers"] = r
        result["sheets_processed"].append("FOLLOWERS")

    if "DEMOGRAPHICS" in sheets:
        r = _import_demographics(wb[sheets["DEMOGRAPHICS"]], batch_id)
        result["demographics"] = r
        result["sheets_processed"].append("DEMOGRAPHICS")

    wb.close()
    logger.info("LinkedIn import complete: %s", result)
    return result


def _import_discovery(ws: object, batch_id: str) -> dict:
    """Parse DISCOVERY sheet: overall impressions + members reached."""
    rows = list(ws.iter_rows(values_only=True))
    data: dict[str, object] = {}
    period = ""

    for row in rows:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().lower()
        if label == "overall performance" and len(row) > 1:
            period = str(row[1]) if row[1] else ""
        elif label == "impressions" and len(row) > 1:
            data["total_impressions"] = _safe_int(row[1])
        elif label == "members reached" and len(row) > 1:
            data["total_members_reached"] = _safe_int(row[1])

    if data:
        conn = get_conn()
        # Parse period for start/end
        period_start, period_end = "", ""
        if " - " in period:
            parts = period.split(" - ")
            period_start = _parse_date(parts[0]) or ""
            period_end = _parse_date(parts[1]) or ""

        conn.execute("DELETE FROM analytics_summary")
        conn.execute(
            """INSERT INTO analytics_summary
               (id, total_impressions, total_members_reached, period_start, period_end, updated_at)
               VALUES (1, ?, ?, ?, ?, datetime('now'))""",
            (data.get("total_impressions", 0), data.get("total_members_reached", 0),
             period_start, period_end),
        )
        conn.commit()

    return {
        "impressions": data.get("total_impressions", 0),
        "members_reached": data.get("total_members_reached", 0),
    }


def _import_engagement(ws: object, batch_id: str) -> dict:
    """Parse ENGAGEMENT sheet: daily impressions + engagements."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"days": 0}

    conn = get_conn()
    count = 0
    for row in rows[1:]:  # Skip header
        if not row or row[0] is None:
            continue
        date = _parse_date(row[0])
        if not date:
            continue
        impressions = _safe_int(row[1]) if len(row) > 1 else 0
        engagements = _safe_int(row[2]) if len(row) > 2 else 0

        conn.execute(
            """INSERT INTO daily_engagement (date, impressions, engagements, import_batch)
               VALUES (?, ?, ?, ?)
               ON CONFLICT(date) DO UPDATE SET
                 impressions = excluded.impressions,
                 engagements = excluded.engagements,
                 import_batch = excluded.import_batch""",
            (date, impressions, engagements, batch_id),
        )
        count += 1

    conn.commit()
    return {"days": count}


def _import_top_posts(ws: object, batch_id: str) -> dict:
    """Parse TOP POSTS sheet: dual-table layout with engagement and impression rankings."""
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row (contains "Post URL")
    header_idx = None
    for i, row in enumerate(rows):
        if row and any(str(c).strip().lower() == "post url" for c in row if c):
            header_idx = i
            break

    if header_idx is None:
        return {"matched": 0, "created": 0, "unmatched": 0}

    # The layout is: [URL, Date, Engagements, None, URL, Date, Impressions, None]
    # Left table: ranked by engagements. Right table: ranked by impressions.
    # Merge them by URL.
    post_data: dict[str, dict] = {}

    for row in rows[header_idx + 1:]:
        if not row:
            continue

        # Left side: engagements
        url_left = str(row[0]).strip().rstrip("/") if row[0] else None
        if url_left and url_left.startswith("http"):
            if url_left not in post_data:
                post_data[url_left] = {"post_url": url_left}
            post_data[url_left]["date_posted"] = _parse_date(row[1]) if len(row) > 1 else None
            post_data[url_left]["engagements"] = _safe_int(row[2]) if len(row) > 2 else 0

        # Right side: impressions
        url_right = str(row[4]).strip().rstrip("/") if len(row) > 4 and row[4] else None
        if url_right and url_right.startswith("http"):
            if url_right not in post_data:
                post_data[url_right] = {"post_url": url_right}
            if not post_data[url_right].get("date_posted"):
                post_data[url_right]["date_posted"] = _parse_date(row[5]) if len(row) > 5 else None
            post_data[url_right]["impressions"] = _safe_int(row[6]) if len(row) > 6 else 0

    if not post_data:
        return {"matched": 0, "created": 0, "unmatched": 0}

    conn = get_conn()
    all_posts = conn.execute("SELECT id, post_url FROM posts").fetchall()

    # Build lookup by activity ID (handles both URL formats)
    activity_id_to_post_id: dict[str, int] = {}
    url_to_id: dict[str, int] = {}
    for p in all_posts:
        if p["post_url"]:
            clean_url = p["post_url"].strip().rstrip("/")
            url_to_id[clean_url] = p["id"]
            aid = _extract_activity_id(clean_url)
            if aid:
                activity_id_to_post_id[aid] = p["id"]

    matched = 0
    created = 0
    now = datetime.now(timezone.utc).isoformat()

    for url, data in post_data.items():
        # Try matching by activity ID first, then by exact URL
        aid = _extract_activity_id(url)
        post_id = None
        if aid:
            post_id = activity_id_to_post_id.get(aid)
        if post_id is None:
            post_id = url_to_id.get(url)

        # Create post if it doesn't exist
        if post_id is None:
            cur = conn.execute(
                """INSERT INTO posts (author, content, post_url, post_type, cta_type,
                   word_count, posted_at, topic_tags)
                   VALUES ('me', '', ?, 'text', 'none', 0, ?, '[]')""",
                (url, data.get("date_posted") or now),
            )
            post_id = cur.lastrowid
            if aid:
                activity_id_to_post_id[aid] = post_id
            created += 1
        else:
            matched += 1

        # XLS only has aggregated engagements + impressions — no breakdown.
        # Strategy: update the latest existing snapshot with better totals,
        # preserving manually entered fields (comments, saves, reposts, etc.).
        # Only create a new snapshot if none exists for this post.
        impressions = data.get("impressions", 0)
        engagements = data.get("engagements", 0)

        latest = conn.execute(
            """SELECT * FROM metrics_snapshots
               WHERE post_id = ? ORDER BY snapshot_at DESC LIMIT 1""",
            (post_id,),
        ).fetchone()

        if latest:
            # Update existing snapshot — only upgrade, never downgrade
            new_imp = max(latest["impressions"] or 0, impressions)
            # XLS "engagements" is total interactions, not just likes.
            # Only update likes if the post has no manual breakdown yet
            # (i.e. comments/saves/reposts are all 0 — meaning no manual entry).
            has_manual_breakdown = (
                (latest["comments"] or 0) > 0
                or (latest["saves"] or 0) > 0
                or (latest["reposts"] or 0) > 0
            )
            if has_manual_breakdown:
                # Keep the manual breakdown, just update impressions if higher
                if new_imp > (latest["impressions"] or 0):
                    new_eng = (
                        (latest["comments"] or 0) * 3
                        + (latest["reposts"] or 0) * 2
                        + (latest["saves"] or 0) * 2
                        + (latest["sends"] or 0) * 1.5
                        + (latest["likes"] or 0)
                    ) / new_imp if new_imp > 0 else 0.0
                    conn.execute(
                        """UPDATE metrics_snapshots
                           SET impressions = ?, engagement_score = ?, snapshot_at = ?
                           WHERE id = ?""",
                        (new_imp, new_eng, now, latest["id"]),
                    )
            else:
                # No manual breakdown — safe to update likes (from engagements) too
                new_likes = max(latest["likes"] or 0, engagements)
                new_eng = (new_likes / new_imp) if new_imp > 0 else 0.0
                conn.execute(
                    """UPDATE metrics_snapshots
                       SET impressions = ?, likes = ?,
                           engagement_score = ?, interaction_score = ?,
                           snapshot_at = ?
                       WHERE id = ?""",
                    (new_imp, new_likes, new_eng, new_likes, now, latest["id"]),
                )
        else:
            # No snapshot exists — create one
            engagement_score = (engagements / impressions) if impressions > 0 else 0.0
            conn.execute(
                """INSERT INTO metrics_snapshots
                   (post_id, impressions, likes, comments, reposts, saves,
                    engagement_score, interaction_score, snapshot_type, snapshot_at)
                   VALUES (?, ?, ?, 0, 0, 0, ?, ?, ?, ?)""",
                (post_id, impressions, engagements, engagement_score, engagements,
                 "linkedin_import", now),
            )

    conn.commit()
    return {"matched": matched, "created": created, "total_posts": len(post_data)}


def _import_followers(ws: object, batch_id: str) -> dict:
    """Parse FOLLOWERS sheet: total + daily new followers."""
    rows = list(ws.iter_rows(values_only=True))
    total_followers = 0
    daily_count = 0

    conn = get_conn()

    for i, row in enumerate(rows):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().lower()

        # First row: "Total followers on 3/7/2026:" with count in col B
        if label.startswith("total followers") and len(row) > 1:
            total_followers = _safe_int(row[1])
            # Update analytics summary
            existing = conn.execute("SELECT id FROM analytics_summary WHERE id = 1").fetchone()
            if existing:
                conn.execute(
                    "UPDATE analytics_summary SET total_followers = ? WHERE id = 1",
                    (total_followers,),
                )
            else:
                conn.execute(
                    """INSERT INTO analytics_summary (id, total_followers) VALUES (1, ?)""",
                    (total_followers,),
                )
            continue

        # Skip header row ("Date", "New followers")
        if label == "date":
            continue

        # Data rows
        date = _parse_date(row[0])
        if not date:
            continue
        new_followers = _safe_int(row[1]) if len(row) > 1 else 0

        conn.execute(
            """INSERT INTO follower_daily (date, new_followers, import_batch)
               VALUES (?, ?, ?)
               ON CONFLICT(date) DO UPDATE SET
                 new_followers = excluded.new_followers,
                 import_batch = excluded.import_batch""",
            (date, new_followers, batch_id),
        )
        daily_count += 1

    conn.commit()
    return {"total_followers": total_followers, "days": daily_count}


def _import_demographics(ws: object, batch_id: str) -> dict:
    """Parse DEMOGRAPHICS sheet: job titles, locations, industries, seniority, company size."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"entries": 0}

    conn = get_conn()
    # Clear old demographics and replace with fresh data
    conn.execute("DELETE FROM audience_demographics")

    count = 0
    for row in rows[1:]:  # Skip header
        if not row or row[0] is None:
            continue
        category = str(row[0]).strip()
        value = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        percentage = _safe_float(row[2]) if len(row) > 2 else 0.0

        if not value:
            continue

        conn.execute(
            """INSERT INTO audience_demographics (category, value, percentage, import_batch)
               VALUES (?, ?, ?, ?)
               ON CONFLICT(category, value) DO UPDATE SET
                 percentage = excluded.percentage,
                 import_batch = excluded.import_batch""",
            (category, value, percentage, batch_id),
        )
        count += 1

    conn.commit()
    return {"entries": count}
